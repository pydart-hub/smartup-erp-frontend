"""
Deep verification: Compare every single data value from the Excel against the JSON.
Reports any mismatches, missing rows, or missing values.
"""
import openpyxl
import json

XLSX = r"FEES STRUCTURE FINAL 2026-27.xlsx"
JSON_FILE = r"fee_structure_2026_27.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
with open(JSON_FILE, encoding="utf-8") as f:
    jdata = json.load(f)

BRANCH_MAP = {
    "ERAVELI": "Eraveli",
    "TIER 1": "Tier 1",
    "THOPPUMPADY": "Thoppumpady",
    "MOOLAMKUZHI": "Moolamkuzhi",
    "VENNALA": "Vennala",
    "KADAVANTHARA": "Kadavanthara",
    "EDAPALLY": "Edapally",
}

errors = []
warnings = []
stats = {"sheets": 0, "sections": 0, "rows_checked": 0, "values_checked": 0}

def find_branch(branch_name):
    for b in jdata["branches"]:
        if b["branch"] == branch_name:
            return b
    return None

def find_category(branch, cat_name):
    for c in branch["fee_categories"]:
        if c["fee_category"] == cat_name:
            return c
    return None

def find_class_entry(category, class_name):
    import re
    cn = re.sub(r'\s+', ' ', class_name.strip())
    for e in category["classes"]:
        if e["class"] == cn:
            return e
    return None

def detect_fee_category(title):
    t = title.strip().lower()
    if "subject wise advanced" in t or "subject wise adv" in t:
        return "Subject-wise Advanced"
    if "subject wise basic" in t:
        return "Subject-wise Basic"
    if "advanced" in t:
        return "Advanced"
    if "intrmediate" in t or "intermediate" in t:
        return "Intermediate"
    if "basic" in t:
        return "Basic"
    return title.strip()

def detect_plan_type(headers):
    header_text = " ".join(str(h or "") for h in headers).lower()
    if "inst plan" in header_text or "inst amount" in header_text:
        return "installment"
    return "quarterly"

def verify_quarterly(entry, headers, row_values, context):
    """Verify quarterly plan values."""
    checks = 0
    for i, h in enumerate(headers):
        if h is None or i >= len(row_values):
            continue
        key = str(h).strip().lower().replace('\n', ' ')
        val = row_values[i]
        if val is None:
            continue

        json_val = None
        field_name = ""

        if "annual" in key:
            json_val = entry.get("annual_fees")
            field_name = "annual_fees"
        elif "early bird" in key:
            json_val = entry.get("payment_plans", {}).get("early_bird")
            field_name = "early_bird"
        elif key in ("otp", "one time payment"):
            json_val = entry.get("payment_plans", {}).get("one_time_payment")
            field_name = "one_time_payment"
        elif "quarterly" in key:
            json_val = entry.get("payment_plans", {}).get("quarterly", {}).get("total_after_5pct_discount")
            field_name = "quarterly_total"
        elif "quarter 1" in key:
            json_val = entry.get("payment_plans", {}).get("quarterly", {}).get("quarter_1")
            field_name = "quarter_1"
        elif "quarter 2" in key:
            json_val = entry.get("payment_plans", {}).get("quarterly", {}).get("quarter_2")
            field_name = "quarter_2"
        elif "quarter 3" in key:
            json_val = entry.get("payment_plans", {}).get("quarterly", {}).get("quarter_3")
            field_name = "quarter_3"
        elif "quarter 4" in key:
            json_val = entry.get("payment_plans", {}).get("quarterly", {}).get("quarter_4")
            field_name = "quarter_4"
        else:
            continue

        checks += 1
        if json_val != val:
            errors.append(f"MISMATCH {context} | {field_name}: Excel={val}, JSON={json_val}")

    return checks

def verify_installment(entry, headers, row_values, context):
    """Verify installment plan values."""
    checks = 0
    for i, h in enumerate(headers):
        if h is None or i >= len(row_values):
            continue
        key = str(h).strip().replace('\n', ' ').lower()
        val = row_values[i]
        if val is None:
            continue

        json_val = None
        field_name = ""

        if "annual" in key:
            json_val = entry.get("annual_fees")
            field_name = "annual_fees"
        elif "early bird" in key:
            json_val = entry.get("payment_plans", {}).get("6_installment", {}).get("early_bird")
            field_name = "6_inst_early_bird"
        elif "6 inst plan" in key:
            json_val = entry.get("payment_plans", {}).get("6_installment", {}).get("total_after_2_5pct_discount")
            field_name = "6_inst_total"
        elif "inst amount" in key and "1st - 5th" in key:
            json_val = entry.get("payment_plans", {}).get("6_installment", {}).get("installment_1_to_5")
            field_name = "6_inst_1to5"
        elif "inst amount" in key and "6th" in key:
            json_val = entry.get("payment_plans", {}).get("6_installment", {}).get("installment_6_final")
            field_name = "6_inst_6final"
        elif "8 inst plan" in key:
            json_val = entry.get("payment_plans", {}).get("8_installment", {}).get("total")
            field_name = "8_inst_total"
        elif "inst amount" in key and "1st - 7th" in key:
            json_val = entry.get("payment_plans", {}).get("8_installment", {}).get("installment_1_to_7")
            field_name = "8_inst_1to7"
        elif "inst amount" in key and "8th" in key:
            json_val = entry.get("payment_plans", {}).get("8_installment", {}).get("installment_8_final")
            field_name = "8_inst_8final"
        else:
            continue

        checks += 1
        if json_val != val:
            errors.append(f"MISMATCH {context} | {field_name}: Excel={val}, JSON={json_val}")

    return checks


for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    branch_name = BRANCH_MAP.get(sheet_name, sheet_name)
    branch = find_branch(branch_name)
    stats["sheets"] += 1

    if not branch:
        errors.append(f"MISSING BRANCH: {branch_name} (sheet: {sheet_name})")
        continue

    # Parse sheet like original
    rows = []
    for r in range(1, ws.max_row + 1):
        row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_data):
            rows.append((r, row_data))

    i = 0
    while i < len(rows):
        row_num, row_data = rows[i]
        first_cell = str(row_data[0] or "").strip()

        if "fees structure" in first_cell.lower():
            category_name = detect_fee_category(first_cell)
            i += 1
            if i >= len(rows):
                break

            _, header_row = rows[i]
            plan_type = detect_plan_type(header_row)
            i += 1
            stats["sections"] += 1

            category = find_category(branch, category_name)
            if not category:
                errors.append(f"MISSING CATEGORY: {branch_name}/{category_name}")
                # Skip data rows
                while i < len(rows):
                    _, dr = rows[i]
                    c0 = str(dr[0] or "").strip().lower()
                    if "fees structure" in c0:
                        break
                    i += 1
                continue

            while i < len(rows):
                _, data_row = rows[i]
                cell0 = str(data_row[0] or "").strip()
                cell0_lower = cell0.lower()
                if "fees structure" in cell0_lower:
                    break
                if not data_row[0] or cell0_lower in ("class", "hss subject"):
                    i += 1
                    continue

                import re
                class_name = re.sub(r'\s+', ' ', cell0.strip())
                entry = find_class_entry(category, class_name)
                context = f"{branch_name}/{category_name}/{class_name} ({plan_type})"

                if not entry:
                    errors.append(f"MISSING CLASS ENTRY: {context}")
                    i += 1
                    stats["rows_checked"] += 1
                    continue

                if plan_type == "quarterly":
                    checks = verify_quarterly(entry, header_row, data_row, context)
                else:
                    checks = verify_installment(entry, header_row, data_row, context)

                stats["rows_checked"] += 1
                stats["values_checked"] += checks
                i += 1
        else:
            i += 1

# Also verify JSON doesn't have extra data not in Excel
print("=" * 70)
print("DEEP VERIFICATION REPORT: Excel vs JSON")
print("=" * 70)
print(f"\nSheets verified:    {stats['sheets']}")
print(f"Sections verified:  {stats['sections']}")
print(f"Data rows verified: {stats['rows_checked']}")
print(f"Values checked:     {stats['values_checked']}")
print()

# Check JSON entry counts per branch
print("--- Entry count per branch ---")
for b in jdata["branches"]:
    total_classes = sum(len(c["classes"]) for c in b["fee_categories"])
    total_cats = len(b["fee_categories"])
    print(f"  {b['branch']:15s}: {total_cats} categories, {total_classes} class entries")

print()

# Verify every JSON entry has non-null values for all expected fields
null_issues = []
for b in jdata["branches"]:
    for cat in b["fee_categories"]:
        for entry in cat["classes"]:
            cn = entry["class"]
            ctx = f"{b['branch']}/{cat['fee_category']}/{cn}"
            if entry.get("annual_fees") is None:
                null_issues.append(f"NULL annual_fees: {ctx}")
            pp = entry.get("payment_plans", {})
            if pp.get("early_bird") is None:
                null_issues.append(f"NULL early_bird: {ctx}")
            if pp.get("one_time_payment") is None:
                null_issues.append(f"NULL one_time_payment: {ctx}")
            q = pp.get("quarterly", {})
            for qf in ["total_after_5pct_discount", "quarter_1", "quarter_2", "quarter_3", "quarter_4"]:
                if q.get(qf) is None:
                    null_issues.append(f"NULL quarterly.{qf}: {ctx}")
            inst6 = pp.get("6_installment", {})
            for f6 in ["early_bird", "total_after_2_5pct_discount", "installment_1_to_5", "installment_6_final"]:
                if inst6.get(f6) is None:
                    null_issues.append(f"NULL 6_installment.{f6}: {ctx}")
            inst8 = pp.get("8_installment", {})
            for f8 in ["total", "installment_1_to_7", "installment_8_final"]:
                if inst8.get(f8) is None:
                    null_issues.append(f"NULL 8_installment.{f8}: {ctx}")

if null_issues:
    print(f"--- NULL value warnings ({len(null_issues)}) ---")
    for ni in null_issues:
        print(f"  ⚠ {ni}")
    print()

if errors:
    print(f"--- ERRORS FOUND ({len(errors)}) ---")
    for e in errors:
        print(f"  ✗ {e}")
else:
    print("✓ ZERO ERRORS: Every single Excel value matches the JSON perfectly!")

if warnings:
    print(f"\n--- Warnings ({len(warnings)}) ---")
    for w in warnings:
        print(f"  ! {w}")

print()
print("=" * 70)
if not errors and not null_issues:
    print("RESULT: PERFECT MATCH - No data loss detected")
elif not errors:
    print(f"RESULT: Values match but {len(null_issues)} null warnings to review")
else:
    print(f"RESULT: {len(errors)} ERRORS need fixing")
print("=" * 70)
