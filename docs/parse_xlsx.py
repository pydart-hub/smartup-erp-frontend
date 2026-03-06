import openpyxl
import json
import csv
import re

xlsx_path = r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE.xlsx'
csv_path = r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE=.csv'

wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb['Sheet1']

# Parse xlsx into structured blocks
blocks = []
current_block = None

for row_idx in range(1, ws.max_row + 1):
    a_val = ws.cell(row=row_idx, column=1).value
    b_val = ws.cell(row=row_idx, column=2).value
    
    if a_val and isinstance(a_val, str) and ('Fees Structure' in a_val or 'fees Structure' in a_val):
        # Header row - extract branch/tier and plan
        header = a_val.strip()
        
        # Determine plan
        plan = None
        if header.lower().startswith('basic'):
            plan = 'Basic'
        elif header.lower().startswith('intrmediate') or header.lower().startswith('intermediate'):
            plan = 'Intermediate'
        elif header.lower().startswith('advanced'):
            plan = 'Advanced'
        
        # Determine branch/tier
        branch = header
        # Try to extract branch name  
        if 'tier 1' in header.lower() or 'chullikal' in header.lower():
            branch = 'Tier 1 (Chullikal, Fortkochi, Eraveli, Palluruthi)'
        elif 'thoppumpady' in header.lower():
            branch = 'Thoppumpady'
        elif 'moolamkuzhi' in header.lower():
            branch = 'Moolamkuzhi'
        elif 'kadavanthra' in header.lower():
            branch = 'Kadavanthra'
        elif 'vennala' in header.lower():
            branch = 'Vennala'
        elif 'edappally' in header.lower():
            branch = 'Edappally'
        
        current_block = {
            'raw_header': header,
            'plan': plan,
            'branch': branch,
            'row_start': row_idx,
            'headers': [],
            'data': []
        }
        blocks.append(current_block)
        
    elif a_val and isinstance(a_val, str) and a_val.strip() == 'Class' and current_block:
        # Column header row
        headers = []
        for col_idx in range(1, 10):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v:
                headers.append(str(v).strip().replace('\n', ' '))
            else:
                headers.append(None)
        current_block['headers'] = headers
        
    elif current_block and a_val and b_val is not None:
        # Data row
        row_data = {}
        for col_idx in range(1, 10):
            v = ws.cell(row=row_idx, column=col_idx).value
            h = current_block['headers'][col_idx-1] if col_idx-1 < len(current_block['headers']) else f'col_{col_idx}'
            if v is not None:
                # Clean up numeric values stored as strings
                if isinstance(v, str):
                    cleaned = v.replace(',', '').replace('₹', '').strip()
                    try:
                        v = int(cleaned)
                    except:
                        try:
                            v = float(cleaned)
                        except:
                            pass
                row_data[h or f'col_{col_idx}'] = v
        if row_data:
            current_block['data'].append(row_data)

# Identify block types
# Blocks alternate: first = quarterly/OTP view, second = instalment view
# Group consecutive blocks of same plan+branch

print("=" * 80)
print("XLSX STRUCTURE ANALYSIS")
print("=" * 80)
print(f"\nTotal blocks found: {len(blocks)}")
print()

for i, b in enumerate(blocks):
    h = b['headers']
    # Determine if it's quarterly or instalment view
    is_quarterly = any('Quarter' in str(x) for x in h if x) or any('OTP' in str(x) for x in h if x) or any('One Time' in str(x) for x in h if x)
    is_instalment = any('Inst Plan' in str(x) for x in h if x) or any('Inst Amount' in str(x) for x in h if x)
    
    view_type = 'QUARTERLY/OTP' if is_quarterly else ('6/8 INSTALMENT' if is_instalment else 'UNKNOWN')
    classes = [d.get('Class', '?') for d in b['data']]
    
    print(f"Block {i+1}: {b['plan']} | {b['branch']} | {view_type} | Classes: {classes}")
    print(f"  Headers: {[x for x in h if x]}")
    print()

print("\n" + "=" * 80)
print("COMPLETE DATA EXTRACTION")
print("=" * 80)

# Now organize into a master data structure
# Key: (branch, plan, class) -> all fee data
master = {}

for i in range(0, len(blocks), 2):
    q_block = blocks[i]  # quarterly/OTP view
    inst_block = blocks[i+1] if i+1 < len(blocks) else None  # instalment view
    
    branch = q_block['branch']
    plan = q_block['plan']
    
    for q_row in q_block['data']:
        cls = str(q_row.get('Class', '')).strip()
        key = (branch, plan, cls)
        
        entry = {
            'branch': branch,
            'plan': plan,
            'class': cls,
            'annual_fee': q_row.get('Annual Fees'),
            'early_bird': q_row.get('Early Bird'),
        }
        
        # OTP - might be labeled "One Time Payment" or "OTP"
        otp = q_row.get('One Time Payment') or q_row.get('OTP')
        entry['otp'] = otp
        
        # Quarterly
        q_total = q_row.get('Quarterly       (5% Disc)')
        entry['quarterly_total'] = q_total
        entry['q1'] = q_row.get('Quarter 1')
        entry['q2'] = q_row.get('Quarter 2')
        entry['q3'] = q_row.get('Quarter 3')
        entry['q4'] = q_row.get('Quarter 4')
        
        # Find matching instalment row
        if inst_block:
            for inst_row in inst_block['data']:
                if str(inst_row.get('Class', '')).strip() == cls:
                    # 6-instalment plan
                    for hk, hv in inst_row.items():
                        if '6 Inst Plan' in str(hk) or '6-Inst Plan' in str(hk):
                            entry['inst6_total'] = hv
                        elif 'Inst Amount' in str(hk) and ('1st - 5th' in str(hk) or '1-5' in str(hk)):
                            entry['inst6_per'] = hv
                        elif 'Inst Amount' in str(hk) and ('6th' in str(hk) or 'LAST' in str(hk)):
                            if 'inst6_last' not in entry:
                                entry['inst6_last'] = hv
                    
                    # 8-instalment plan
                    for hk, hv in inst_row.items():
                        if '8 Inst Plan' in str(hk) or '8-Inst Plan' in str(hk):
                            entry['inst8_total'] = hv
                        elif 'Inst Amount' in str(hk) and ('1st - 7th' in str(hk) or '1-7' in str(hk)):
                            entry['inst8_per'] = hv
                        elif '8th' in str(hk) or (entry.get('inst6_last') and 'LAST' in str(hk)):
                            if hk != list(inst_row.keys())[-2]:  # not the 6th last
                                pass
                    
                    # More precise extraction from instalment block
                    h = inst_block['headers']
                    vals = list(inst_row.values())
                    # Col D (index 3) = 6-inst total
                    # Col E (index 4) = 6-inst per (1st-5th)  
                    # Col F (index 5) = 6-inst last (6th)
                    # Col G (index 6) = 8-inst total
                    # Col H (index 7) = 8-inst per (1st-7th)
                    # Col I (index 8) = 8-inst last (8th)
                    raw_vals = []
                    for col_idx in range(1, 10):
                        v = None
                        for r in range(inst_block['row_start'], inst_block['row_start'] + 20):
                            cell_a = ws.cell(row=r, column=1).value
                            if cell_a and str(cell_a).strip() == cls:
                                v = ws.cell(row=r, column=col_idx).value
                                if isinstance(v, str):
                                    cleaned = v.replace(',', '').replace('₹', '').strip()
                                    try:
                                        v = int(cleaned)
                                    except:
                                        try:
                                            v = float(cleaned)
                                        except:
                                            pass
                                break
                        raw_vals.append(v)
                    
                    if len(raw_vals) >= 9:
                        entry['inst6_total'] = raw_vals[3]  # D
                        entry['inst6_per'] = raw_vals[4]    # E
                        entry['inst6_last'] = raw_vals[5]   # F
                        entry['inst8_total'] = raw_vals[6]  # G
                        entry['inst8_per'] = raw_vals[7]    # H 
                        entry['inst8_last'] = raw_vals[8]   # I
                    
                    break
        
        master[key] = entry

# Print organized output
print(f"\nTotal unique fee entries: {len(master)}")

branches_order = [
    'Tier 1 (Chullikal, Fortkochi, Eraveli, Palluruthi)',
    'Thoppumpady',
    'Moolamkuzhi',
    'Kadavanthra',
    'Vennala',
    'Edappally'
]
plans_order = ['Basic', 'Intermediate', 'Advanced']

for branch in branches_order:
    print(f"\n{'='*80}")
    print(f"BRANCH: {branch}")
    print(f"{'='*80}")
    
    for plan in plans_order:
        entries = [(k, v) for k, v in master.items() if k[0] == branch and k[1] == plan]
        if not entries:
            continue
        
        print(f"\n  --- {plan} ---")
        print(f"  {'Class':<12} {'Annual':>8} {'EarlyBird':>10} {'OTP':>8} {'QTotal':>8} {'Q1':>6} {'Q2':>6} {'Q3':>6} {'Q4':>6} | {'6I-Tot':>8} {'6I-Per':>7} {'6I-Last':>8} {'8I-Tot':>8} {'8I-Per':>7} {'8I-Last':>8}")
        print(f"  {'-'*12} {'-'*8} {'-'*10} {'-'*8} {'-'*8} {'-'*6} {'-'*6} {'-'*6} {'-'*6} | {'-'*8} {'-'*7} {'-'*8} {'-'*8} {'-'*7} {'-'*8}")
        
        for k, v in sorted(entries, key=lambda x: x[0][2]):
            cls = v['class']
            print(f"  {cls:<12} {v.get('annual_fee','?'):>8} {v.get('early_bird','?'):>10} {v.get('otp','?'):>8} {v.get('quarterly_total','?'):>8} {v.get('q1','?'):>6} {v.get('q2','?'):>6} {v.get('q3','?'):>6} {v.get('q4','?'):>6} | {v.get('inst6_total','?'):>8} {v.get('inst6_per','?'):>7} {v.get('inst6_last','?'):>8} {v.get('inst8_total','?'):>8} {v.get('inst8_per','?'):>7} {v.get('inst8_last','?'):>8}")

# Validation checks
print(f"\n\n{'='*80}")
print("VALIDATION CHECKS")
print(f"{'='*80}")

for key, v in sorted(master.items()):
    branch, plan, cls = key
    errors = []
    
    # Q1+Q2+Q3+Q4 should equal quarterly_total
    if all(v.get(x) for x in ['q1','q2','q3','q4','quarterly_total']):
        q_sum = v['q1'] + v['q2'] + v['q3'] + v['q4']
        if q_sum != v['quarterly_total']:
            errors.append(f"Q-sum={q_sum} != Q-total={v['quarterly_total']} (diff={q_sum - v['quarterly_total']})")
    
    # 6-inst: 5*per + last should equal inst6_total
    if all(v.get(x) for x in ['inst6_per','inst6_last','inst6_total']):
        calc = 5 * v['inst6_per'] + v['inst6_last']
        if calc != v['inst6_total']:
            errors.append(f"6I: 5*{v['inst6_per']}+{v['inst6_last']}={calc} != total={v['inst6_total']} (diff={calc - v['inst6_total']})")
    
    # 8-inst: 7*per + last should equal inst8_total
    if all(v.get(x) for x in ['inst8_per','inst8_last','inst8_total']):
        calc = 7 * v['inst8_per'] + v['inst8_last']
        if calc != v['inst8_total']:
            errors.append(f"8I: 7*{v['inst8_per']}+{v['inst8_last']}={calc} != total={v['inst8_total']} (diff={calc - v['inst8_total']})")
    
    # OTP should be less than early_bird
    if v.get('otp') and v.get('early_bird'):
        if v['otp'] > v['early_bird']:
            errors.append(f"OTP={v['otp']} > EarlyBird={v['early_bird']} (REVERSED!)")
    
    # Early bird should be less than annual
    if v.get('early_bird') and v.get('annual_fee'):
        if v['early_bird'] > v['annual_fee']:
            errors.append(f"EarlyBird={v['early_bird']} > Annual={v['annual_fee']} (REVERSED!)")
    
    # 6-inst total should be ~97.5% of early bird 
    if v.get('inst6_total') and v.get('early_bird'):
        expected = v['early_bird'] * 0.975
        diff_pct = abs(v['inst6_total'] - expected) / expected * 100
        if diff_pct > 1:
            errors.append(f"6I-total={v['inst6_total']} vs 97.5%*EB={expected:.0f} (diff={diff_pct:.1f}%)")
    
    if errors:
        print(f"\n  {branch} | {plan} | {cls}:")
        for e in errors:
            print(f"    - {e}")

# Summary stats
print(f"\n\n{'='*80}")
print("SUMMARY STATISTICS")
print(f"{'='*80}")

branch_class_map = {}
for key in master:
    branch, plan, cls = key
    if branch not in branch_class_map:
        branch_class_map[branch] = set()
    branch_class_map[branch].add(cls)

print("\nClasses available per branch:")
for branch in branches_order:
    classes = sorted(branch_class_map.get(branch, set()))
    print(f"  {branch}: {classes}")

print(f"\nTotal unique (branch, plan, class) combos: {len(master)}")

# Count by payment option (each combo has 4 options: OTP, Quarterly, 6-inst, 8-inst)
# So total fee structure records needed = combos * 4
print(f"Total fee structure records needed (combos × 4 payment options): {len(master) * 4}")

# Write to JSON for reference
output = {}
for key, v in master.items():
    branch, plan, cls = key
    k = f"{branch}|{plan}|{cls}"
    output[k] = v

with open(r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\fee_structure_parsed.json', 'w') as f:
    json.dump(output, f, indent=2, default=str)

print(f"\nParsed data written to docs/fee_structure_parsed.json")
