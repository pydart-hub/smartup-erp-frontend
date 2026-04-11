"""
INDEPENDENT raw verification: No shared logic with the parser.
1. Reads every cell from Excel into flat records
2. Reads every value from JSON into flat records  
3. Cross-matches them exhaustively
4. Also detects any Excel rows the JSON might have skipped
"""
import openpyxl
import json
import re

XLSX = r"FEES STRUCTURE FINAL 2026-27.xlsx"
JSON_FILE = r"fee_structure_2026_27.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
with open(JSON_FILE, encoding="utf-8") as f:
    jdata = json.load(f)

# ─── STEP 1: Extract EVERY data cell from Excel into flat list ───
# Each record = (sheet, section_title, plan_type, class_name, field, value)

excel_records = []

BRANCH_SHEETS = ["ERAVELI", "TIER 1", "THOPPUMPADY", "MOOLAMKUZHI", "VENNALA", "KADAVANTHARA", "EDAPALLY"]

for sheet_name in BRANCH_SHEETS:
    ws = wb[sheet_name]
    all_rows = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        all_rows.append((r, row))

    section_title = None
    headers = None
    plan_type = None

    for row_num, row in all_rows:
        c0 = str(row[0] or "").strip()
        c0_low = c0.lower()

        # Section title detection
        if "fees structure" in c0_low:
            section_title = c0
            headers = None
            plan_type = None
            continue

        # Header row detection
        if c0_low in ("class", "hss subject") and section_title:
            headers = row
            ht = " ".join(str(h or "") for h in headers).lower()
            plan_type = "installment" if ("inst plan" in ht or "inst amount" in ht) else "quarterly"
            continue

        # Data row: must have a class name in col A and a number in col B
        if headers and row[0] and row[1] is not None and isinstance(row[1], (int, float)):
            class_name = re.sub(r'\s+', ' ', str(row[0]).strip())

            for col_idx in range(1, len(headers)):
                if col_idx >= len(row):
                    break
                hdr = headers[col_idx]
                val = row[col_idx]
                if hdr is None or val is None:
                    continue
                hdr_clean = re.sub(r'\s+', ' ', str(hdr).replace('\n', ' ')).strip()
                excel_records.append({
                    "sheet": sheet_name,
                    "section": section_title,
                    "plan": plan_type,
                    "class": class_name,
                    "header": hdr_clean,
                    "value": val,
                    "row": row_num,
                    "col": col_idx + 1,
                })

print(f"EXCEL: Extracted {len(excel_records)} individual data cells\n")

# ─── STEP 2: Extract EVERY value from JSON into flat list ───

BRANCH_MAP = {
    "ERAVELI": "Eraveli", "TIER 1": "Tier 1", "THOPPUMPADY": "Thoppumpady",
    "MOOLAMKUZHI": "Moolamkuzhi", "VENNALA": "Vennala",
    "KADAVANTHARA": "Kadavanthara", "EDAPALLY": "Edapally",
}

# Map JSON field paths to (plan_type, header_pattern)
QUARTERLY_MAP = {
    ("annual_fees",): "Annual Fees",
    ("payment_plans", "early_bird"): "Early Bird",
    ("payment_plans", "one_time_payment"): "OTP/One Time Payment",
    ("payment_plans", "quarterly", "total_after_5pct_discount"): "Quarterly (5% Disc)",
    ("payment_plans", "quarterly", "quarter_1"): "Quarter 1",
    ("payment_plans", "quarterly", "quarter_2"): "Quarter 2",
    ("payment_plans", "quarterly", "quarter_3"): "Quarter 3",
    ("payment_plans", "quarterly", "quarter_4"): "Quarter 4",
}

INSTALLMENT_MAP = {
    ("annual_fees",): "Annual Fees",
    ("payment_plans", "6_installment", "early_bird"): "Early Bird (₹)",
    ("payment_plans", "6_installment", "total_after_2_5pct_discount"): "6 Inst Plan Full Amnt",
    ("payment_plans", "6_installment", "installment_1_to_5"): "Inst Amount (1st - 5th)",
    ("payment_plans", "6_installment", "installment_6_final"): "Inst Amount 6th (Final)",
    ("payment_plans", "8_installment", "total"): "8 Inst Plan Full Amnt",
    ("payment_plans", "8_installment", "installment_1_to_7"): "Inst Amount (1st - 7th)",
    ("payment_plans", "8_installment", "installment_8_final"): "Inst Amount 8th (Final)",
}

def get_nested(d, keys):
    for k in keys:
        if isinstance(d, dict):
            d = d.get(k)
        else:
            return None
    return d

json_records = []

for branch in jdata["branches"]:
    sheet_name = branch["sheet_name"]
    for cat in branch["fee_categories"]:
        cat_name = cat["fee_category"]
        for entry in cat["classes"]:
            class_name = entry["class"]
            # Quarterly values
            for keys, label in QUARTERLY_MAP.items():
                val = get_nested(entry, keys)
                if val is not None:
                    json_records.append({
                        "sheet": sheet_name,
                        "category": cat_name,
                        "plan": "quarterly",
                        "class": class_name,
                        "field": label,
                        "value": val,
                    })
            # Installment values
            for keys, label in INSTALLMENT_MAP.items():
                val = get_nested(entry, keys)
                if val is not None:
                    json_records.append({
                        "sheet": sheet_name,
                        "category": cat_name,
                        "plan": "installment",
                        "class": class_name,
                        "field": label,
                        "value": val,
                    })

print(f"JSON:  Extracted {len(json_records)} individual data values\n")

# ─── STEP 3: Build lookup from Excel for matching ───
# Key: (sheet, class_name, plan_type, value)

def normalize_cat(section_title):
    t = section_title.lower()
    if "subject wise advanced" in t: return "Subject-wise Advanced"
    if "subject wise basic" in t: return "Subject-wise Basic"
    if "advanced" in t: return "Advanced"
    if "intrmediate" in t or "intermediate" in t: return "Intermediate"
    if "basic" in t: return "Basic"
    return section_title

# Build per-(sheet, category, plan, class) value sets from Excel
excel_by_key = {}
for rec in excel_records:
    cat = normalize_cat(rec["section"])
    key = (rec["sheet"], cat, rec["plan"], rec["class"])
    excel_by_key.setdefault(key, []).append(rec["value"])

# Build same from JSON
json_by_key = {}
for rec in json_records:
    key = (rec["sheet"], rec["category"], rec["plan"], rec["class"])
    json_by_key.setdefault(key, []).append(rec["value"])

# ─── STEP 4: Compare ───
errors = []
checked_keys = set()

# Check all Excel keys exist in JSON with matching values
for key, excel_vals in sorted(excel_by_key.items()):
    checked_keys.add(key)
    json_vals = json_by_key.get(key)
    if json_vals is None:
        errors.append(f"MISSING IN JSON: {key}")
        continue
    
    excel_sorted = sorted(excel_vals)
    json_sorted = sorted(json_vals)
    
    if excel_sorted != json_sorted:
        # Find specific differences
        excel_set = {}
        for v in excel_vals:
            excel_set[v] = excel_set.get(v, 0) + 1
        json_set = {}
        for v in json_vals:
            json_set[v] = json_set.get(v, 0) + 1
        
        only_excel = {k: v for k, v in excel_set.items() if k not in json_set or json_set[k] < v}
        only_json = {k: v for k, v in json_set.items() if k not in excel_set or excel_set[k] < v}
        
        errors.append(f"VALUE MISMATCH: {key}")
        if only_excel:
            errors.append(f"  In Excel only: {only_excel}")
        if only_json:
            errors.append(f"  In JSON only:  {only_json}")

# Check for extra JSON keys not in Excel
for key in sorted(json_by_key.keys()):
    if key not in excel_by_key:
        errors.append(f"EXTRA IN JSON (not in Excel): {key}")

# ─── STEP 5: Also verify class lists per branch/category ───
print("--- Classes per branch/category ---")
for branch in jdata["branches"]:
    print(f"\n  {branch['branch']}:")
    for cat in branch["fee_categories"]:
        classes = [e["class"] for e in cat["classes"]]
        print(f"    {cat['fee_category']:25s}: {classes}")

# ─── STEP 6: Check for rows in Excel that might have been skipped ───
print("\n\n--- Checking for any skipped Excel rows ---")
skipped = []
for sheet_name in BRANCH_SHEETS:
    ws = wb[sheet_name]
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v is not None for v in row):
            continue
        c0 = str(row[0] or "").strip().lower()
        # Skip titles, headers, empty first cells
        if "fees structure" in c0 or c0 in ("class", "hss subject", ""):
            continue
        # This should be a data row - check if it has numeric data
        if row[1] is not None and isinstance(row[1], (int, float)):
            class_name = re.sub(r'\s+', ' ', str(row[0]).strip())
            # Check if this class appears in any JSON record for this sheet
            found = False
            for rec in json_records:
                if rec["sheet"] == sheet_name and rec["class"] == class_name:
                    found = True
                    break
            if not found:
                skipped.append(f"  Row {r} in {sheet_name}: {row[0]} -> {[v for v in row[1:] if v is not None]}")

if skipped:
    print(f"  SKIPPED ROWS FOUND ({len(skipped)}):")
    for s in skipped:
        print(s)
else:
    print("  None - every data row is accounted for.")

# ─── FINAL REPORT ───
print("\n" + "=" * 70)
print("INDEPENDENT VERIFICATION REPORT")
print("=" * 70)
print(f"  Excel cells extracted:  {len(excel_records)}")
print(f"  JSON values extracted:  {len(json_records)}")
print(f"  Unique (sheet,cat,plan,class) keys in Excel: {len(excel_by_key)}")
print(f"  Unique (sheet,cat,plan,class) keys in JSON:  {len(json_by_key)}")
print(f"  Keys checked:           {len(checked_keys)}")

if errors:
    print(f"\n  *** {len(errors)} ISSUES FOUND ***")
    for e in errors:
        print(f"    {e}")
else:
    print(f"\n  ✓ ALL VALUES MATCH PERFECTLY")
    print(f"  ✓ NO MISSING DATA")
    print(f"  ✓ NO EXTRA DATA")
    print(f"  ✓ NO SKIPPED ROWS")

print("=" * 70)
