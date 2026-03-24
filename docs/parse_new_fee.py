"""
Parse 'New Fee.xlsx' into fee_structure_parsed JSON format.

Uses per-branch sheets (Tier 1, Thoppumpady, Moolamkuzhi, Kadavanthra, Vennala, Edappally)
as the authoritative source. Each sheet has 4 sections:
  1. Basic OTP & Quarterly (cols A-I)
  2. Basic 6-inst & 8-inst (cols A-I, separate block)
  3. Advanced OTP & Quarterly (cols A-I)
  4. Advanced 6-inst & 8-inst (cols A-I, separate block)
Plus optional "Subject wise" sections for HSS subjects at Kadavanthra, Vennala, Edappally.

Output format matches the existing fee_structure_parsed.json schema:
  Key: "{branch}|{plan}|{class}"
  Value: { branch, plan, class, annual_fee, early_bird, otp, quarterly_total,
           q1, q2, q3, q4, inst6_total, inst6_per, inst6_last,
           inst8_total, inst8_per, inst8_last }
"""

import json
import openpyxl
import sys
from pathlib import Path

XLSX_PATH = Path(__file__).parent / "New Fee.xlsx"
OUTPUT_PATH = Path(__file__).parent / "fee_structure_parsed copy.json"

# Branch name mapping (sheet name → JSON branch key)
BRANCH_MAP = {
    "Tier 1": "Tier 1 (Chullikal, Fortkochi, Eraveli, Palluruthi)",
    "Thoppumpady": "Thoppumpady",
    "Moolamkuzhi": "Moolamkuzhi",
    "Kadavanthra": "Kadavanthra",
    "Vennala": "Vennala",
    "Edappally": "Edappally",
}

SHEETS_TO_PARSE = list(BRANCH_MAP.keys())


def find_sections(ws):
    """
    Scan a worksheet and identify all data sections.
    Returns a list of dicts:
      { 'plan': 'Basic'|'Advanced',
        'type': 'otp_qtr' | 'inst6_8' | 'subject_otp_qtr' | 'subject_inst6_8',
        'header_row': int,
        'data_start': int,
        'data_end': int }
    """
    sections = []
    max_row = ws.max_row

    for r in range(1, max_row + 1):
        cell_a = ws.cell(row=r, column=1).value
        if not cell_a or not isinstance(cell_a, str):
            continue

        # Detect section title rows
        cell_lower = cell_a.strip().lower()

        plan = None
        section_type = None

        if "basic fees structure" in cell_lower or "basic fee structure" in cell_lower:
            plan = "Basic"
        elif "advanced fees structure" in cell_lower or "advanced fee structure" in cell_lower:
            plan = "Advanced"
        elif "subject wise basic" in cell_lower:
            plan = "Basic"
            section_type = "subject"
        elif "subject wise advanced" in cell_lower:
            plan = "Advanced"
            section_type = "subject"

        if plan is None:
            continue

        # Find header row (next row with "Class" or "HSS Subject" in column A)
        header_row = None
        for hr in range(r + 1, min(r + 5, max_row + 1)):
            hv = ws.cell(row=hr, column=1).value
            if hv and isinstance(hv, str) and hv.strip().lower() in ("class", "hss subject"):
                header_row = hr
                break

        if header_row is None:
            continue

        # Determine if this is OTP/QTR or 6/8-inst by looking at header row columns
        header_d = ws.cell(row=header_row, column=4).value
        if header_d and isinstance(header_d, str):
            header_d_clean = header_d.strip().lower()
        else:
            header_d_clean = ""

        if "6 inst" in header_d_clean or "6-inst" in header_d_clean or "after disc" in header_d_clean:
            data_type = "inst6_8"
        else:
            data_type = "otp_qtr"

        if section_type == "subject":
            data_type = "subject_" + data_type

        # Collect data rows (class name in col A, numeric in col B)
        data_start = header_row + 1
        data_end = data_start
        for dr in range(data_start, max_row + 1):
            av = ws.cell(row=dr, column=1).value
            bv = ws.cell(row=dr, column=2).value
            if av and bv and isinstance(bv, (int, float)):
                data_end = dr
            elif av and isinstance(av, str) and av.strip().lower() in ("class", "hss subject"):
                break
            elif av and isinstance(av, str) and ("fees structure" in av.lower() or "fee structure" in av.lower()):
                break
            elif av is None and bv is None:
                # Could be empty gap row — check if next rows have more data
                # Look ahead 2 rows
                next_av = ws.cell(row=dr + 1, column=1).value if dr + 1 <= max_row else None
                next_bv = ws.cell(row=dr + 1, column=2).value if dr + 1 <= max_row else None
                if next_av and next_bv and isinstance(next_bv, (int, float)):
                    continue  # gap row, skip
                else:
                    break

        sections.append({
            "plan": plan,
            "type": data_type,
            "header_row": header_row,
            "data_start": data_start,
            "data_end": data_end,
        })

    return sections


def parse_otp_qtr_rows(ws, start_row, end_row):
    """
    Parse OTP & Quarterly section rows.
    Columns: A=Class, B=Annual, C=EarlyBird, D=OTP, E=QuarterlyTotal, F=Q1, G=Q2, H=Q3, I=Q4
    Returns list of dicts with partial data.
    """
    rows = []
    for r in range(start_row, end_row + 1):
        cls = ws.cell(row=r, column=1).value
        annual = ws.cell(row=r, column=2).value
        if not cls or not annual or not isinstance(annual, (int, float)):
            continue
        cls_name = cls.strip()
        rows.append({
            "class": cls_name,
            "annual_fee": int(annual),
            "early_bird": int(ws.cell(row=r, column=3).value or 0),
            "otp": int(ws.cell(row=r, column=4).value or 0),
            "quarterly_total": int(ws.cell(row=r, column=5).value or 0),
            "q1": int(ws.cell(row=r, column=6).value or 0),
            "q2": int(ws.cell(row=r, column=7).value or 0),
            "q3": int(ws.cell(row=r, column=8).value or 0),
            "q4": int(ws.cell(row=r, column=9).value or 0),
        })
    return rows


def parse_inst6_8_rows(ws, start_row, end_row):
    """
    Parse 6-inst & 8-inst section rows.
    Columns: A=Class, B=Annual, C=EarlyBird, D=6InstTotal, E=Inst1-5, F=Inst6(last),
             G=8InstTotal, H=Inst1-7, I=Inst8(last)
    Returns list of dicts with partial data.
    """
    rows = []
    for r in range(start_row, end_row + 1):
        cls = ws.cell(row=r, column=1).value
        annual = ws.cell(row=r, column=2).value
        if not cls or not annual or not isinstance(annual, (int, float)):
            continue
        cls_name = cls.strip()
        rows.append({
            "class": cls_name,
            "annual_fee": int(annual),
            "early_bird": int(ws.cell(row=r, column=3).value or 0),
            "inst6_total": int(ws.cell(row=r, column=4).value or 0),
            "inst6_per": int(ws.cell(row=r, column=5).value or 0),
            "inst6_last": int(ws.cell(row=r, column=6).value or 0),
            "inst8_total": int(ws.cell(row=r, column=7).value or 0),
            "inst8_per": int(ws.cell(row=r, column=8).value or 0),
            "inst8_last": int(ws.cell(row=r, column=9).value or 0),
        })
    return rows


def normalize_class_name(name):
    """Normalize class names for consistency."""
    name = name.strip()
    # Normalize spacing in compound names
    name = name.replace("Plus One ", "Plus One").replace("Plus Two ", "Plus Two")
    name = name.replace("Phy - Chem", "Phy-Chem")
    name = name.replace("Chem- maths", "Chem-Maths").replace("Chem- Maths", "Chem-Maths")
    name = name.replace("Chemitry", "Chemistry")
    name = name.strip()
    return name


def parse_sheet(ws, branch_name):
    """Parse a single branch sheet and return a dict of fee entries."""
    sections = find_sections(ws)
    entries = {}  # key: (plan, class) → merged dict

    for sec in sections:
        is_subject = sec["type"].startswith("subject_")
        base_type = sec["type"].replace("subject_", "")

        if base_type == "otp_qtr":
            rows = parse_otp_qtr_rows(ws, sec["data_start"], sec["data_end"])
            for row in rows:
                cls = normalize_class_name(row["class"])
                key = (sec["plan"], cls)
                if key not in entries:
                    entries[key] = {"branch": branch_name, "plan": sec["plan"], "class": cls}
                entries[key].update({
                    "annual_fee": row["annual_fee"],
                    "early_bird": row["early_bird"],
                    "otp": row["otp"],
                    "quarterly_total": row["quarterly_total"],
                    "q1": row["q1"],
                    "q2": row["q2"],
                    "q3": row["q3"],
                    "q4": row["q4"],
                })

        elif base_type == "inst6_8":
            rows = parse_inst6_8_rows(ws, sec["data_start"], sec["data_end"])
            for row in rows:
                cls = normalize_class_name(row["class"])
                key = (sec["plan"], cls)
                if key not in entries:
                    entries[key] = {"branch": branch_name, "plan": sec["plan"], "class": cls}
                entries[key].update({
                    "inst6_total": row["inst6_total"],
                    "inst6_per": row["inst6_per"],
                    "inst6_last": row["inst6_last"],
                    "inst8_total": row["inst8_total"],
                    "inst8_per": row["inst8_per"],
                    "inst8_last": row["inst8_last"],
                })
                # Also set annual_fee and early_bird if not already set
                if "annual_fee" not in entries[key]:
                    entries[key]["annual_fee"] = row["annual_fee"]
                    entries[key]["early_bird"] = row["early_bird"]

    return entries


def validate_entry(key_str, entry):
    """Check that all 17 fields are present and non-negative."""
    required = [
        "branch", "plan", "class",
        "annual_fee", "early_bird", "otp", "quarterly_total",
        "q1", "q2", "q3", "q4",
        "inst6_total", "inst6_per", "inst6_last",
        "inst8_total", "inst8_per", "inst8_last",
    ]
    missing = [f for f in required if f not in entry]
    if missing:
        print(f"  WARNING: {key_str} missing fields: {missing}")
        return False
    return True


def main():
    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)
    all_entries = {}

    for sheet_name in SHEETS_TO_PARSE:
        branch_name = BRANCH_MAP[sheet_name]
        ws = wb[sheet_name]
        print(f"\nParsing sheet: {sheet_name} → {branch_name}")

        entries = parse_sheet(ws, branch_name)

        for (plan, cls), entry in sorted(entries.items()):
            json_key = f"{branch_name}|{plan}|{cls}"
            # Ensure all fields present with defaults
            final = {
                "branch": branch_name,
                "plan": plan,
                "class": cls,
                "annual_fee": entry.get("annual_fee", 0),
                "early_bird": entry.get("early_bird", 0),
                "otp": entry.get("otp", 0),
                "quarterly_total": entry.get("quarterly_total", 0),
                "q1": entry.get("q1", 0),
                "q2": entry.get("q2", 0),
                "q3": entry.get("q3", 0),
                "q4": entry.get("q4", 0),
                "inst6_total": entry.get("inst6_total", 0),
                "inst6_per": entry.get("inst6_per", 0),
                "inst6_last": entry.get("inst6_last", 0),
                "inst8_total": entry.get("inst8_total", 0),
                "inst8_per": entry.get("inst8_per", 0),
                "inst8_last": entry.get("inst8_last", 0),
            }
            validate_entry(json_key, entry)
            all_entries[json_key] = final
            print(f"  {json_key}")

    # Summary
    print(f"\n{'='*60}")
    print(f"Total entries: {len(all_entries)}")

    branches = sorted(set(v["branch"] for v in all_entries.values()))
    plans = sorted(set(v["plan"] for v in all_entries.values()))
    classes = sorted(set(v["class"] for v in all_entries.values()))
    print(f"Branches ({len(branches)}): {branches}")
    print(f"Plans ({len(plans)}): {plans}")
    print(f"Classes ({len(classes)}): {classes}")

    # Matrix: branch × plan × class
    print(f"\nCoverage matrix:")
    for branch in branches:
        short = branch[:20]
        for plan in plans:
            cls_list = [v["class"] for k, v in all_entries.items()
                        if v["branch"] == branch and v["plan"] == plan]
            print(f"  {short:20s} | {plan:10s} | {len(cls_list):2d} classes: {', '.join(sorted(cls_list))}")

    # Write output
    with open(str(OUTPUT_PATH), "w", encoding="utf-8") as f:
        json.dump(all_entries, f, indent=2, ensure_ascii=False)
    print(f"\nWritten to: {OUTPUT_PATH}")
    print(f"Total entries: {len(all_entries)}")


if __name__ == "__main__":
    main()
