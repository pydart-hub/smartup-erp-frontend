import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE.xlsx', data_only=True)
print(f'Sheet names: {wb.sheetnames}')
print('---')
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n===== SHEET: {sname} =====')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f'{cell.coordinate}={v}')
        if vals:
            print(' | '.join(vals))
        else:
            print('---empty---')
