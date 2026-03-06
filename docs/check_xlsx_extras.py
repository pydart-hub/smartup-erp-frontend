import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE.xlsx')
ws = wb['Sheet1']

# Check for merged cells
print("MERGED CELLS:", list(ws.merged_cells.ranges))

# Check for comments/notes
comments = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.comment:
            comments.append(f"{cell.coordinate}: {cell.comment.text}")
print(f"\nCOMMENTS: {comments if comments else 'None'}")

# Check for colored cells or special formatting
colored = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb not in ('00000000', '0', None):
            if cell.value is not None:
                colored.append(f"{cell.coordinate}={cell.value} (color={fill.fgColor.rgb})")

print(f"\nCOLORED CELLS (first 50): {colored[:50] if colored else 'None'}")
print(f"Total colored cells: {len(colored)}")

# Check for formulas (load without data_only)
formula_cells = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_cells.append(f"{cell.coordinate}: {cell.value}")
print(f"\nFORMULA CELLS: {formula_cells if formula_cells else 'None'}")

# Check for hidden rows/columns
hidden_rows = [r for r in range(1, ws.max_row+1) if ws.row_dimensions[r].hidden]
hidden_cols = [c for c in range(1, ws.max_column+1) if ws.column_dimensions[openpyxl.utils.get_column_letter(c)].hidden]
print(f"\nHIDDEN ROWS: {hidden_rows if hidden_rows else 'None'}")
print(f"HIDDEN COLS: {hidden_cols if hidden_cols else 'None'}")

# Check for data validation (dropdowns etc)
print(f"\nDATA VALIDATIONS: {list(ws.data_validations.dataValidation) if ws.data_validations else 'None'}")

# Look for any text that contains keywords we might have missed
interesting_text = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            v = cell.value.strip().lower()
            keywords = ['discount', 'scholarship', 'sibling', 'penalty', 'late', 'fine',
                       'gst', 'tax', 'registration', 'admission', 'exam', 'book',
                       'transport', 'hostel', 'uniform', 'material', 'lab', 'library',
                       'sport', 'extra', 'note', 'important', 'condition', 'terms',
                       'refund', 'cancellation', 'transfer', 'concession', 'waiver']
            for kw in keywords:
                if kw in v:
                    interesting_text.append(f"{cell.coordinate}: {cell.value}")
                    break

print(f"\nINTERESTING TEXT (keywords): {interesting_text if interesting_text else 'None'}")

# Check for any data beyond column I
extra_cols = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=10, max_col=20):
    for cell in row:
        if cell.value is not None:
            extra_cols.append(f"{cell.coordinate}={cell.value}")
print(f"\nDATA BEYOND COLUMN I: {extra_cols if extra_cols else 'None'}")

# Check if the xlsx has formulas that resolve differently
wb2 = openpyxl.load_workbook(r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE.xlsx')
ws2 = wb2['Sheet1']
diffs = []
for row in range(1, ws.max_row+1):
    for col in range(1, 10):
        v_formula = ws2.cell(row=row, column=col).value
        if v_formula and isinstance(v_formula, str) and v_formula.startswith('='):
            diffs.append(f"Row {row}, Col {col}: formula={v_formula}")
print(f"\nFORMULAS IN XLSX: {diffs[:20] if diffs else 'None'}")

# Compare xlsx vs csv - look for discrepancies  
import csv
csv_path = r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE=.csv'
csv_data = []
with open(csv_path, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row in reader:
        csv_data.append(row)

print(f"\nCSV rows: {len(csv_data)}")
print(f"XLSX rows: {ws.max_row}")

# Detailed diff
wb_data = openpyxl.load_workbook(r'C:\Users\arjun\Desktop\Pydart\smartup-erp-frontend\docs\FINAL FEE STRUCTURE.xlsx', data_only=True)
ws_data = wb_data['Sheet1']

mismatches = []
for row_idx in range(1, min(ws_data.max_row, len(csv_data)) + 1):
    for col_idx in range(1, 10):
        xlsx_val = ws_data.cell(row=row_idx, column=col_idx).value
        csv_col = col_idx - 1
        csv_row = row_idx - 1
        
        if csv_row < len(csv_data) and csv_col < len(csv_data[csv_row]):
            csv_val = csv_data[csv_row][csv_col].strip() if csv_data[csv_row][csv_col] else None
        else:
            csv_val = None
        
        # Compare
        if xlsx_val is None and (csv_val is None or csv_val == ''):
            continue
        
        if xlsx_val is not None and csv_val is not None:
            # Try numeric comparison
            try:
                csv_num = float(csv_val.replace(',', '').replace('"', '').replace('₹', ''))
                xlsx_num = float(xlsx_val) if not isinstance(xlsx_val, str) else float(xlsx_val.replace(',', ''))
                if abs(csv_num - xlsx_num) > 0.01:
                    mismatches.append(f"Row {row_idx}, Col {col_idx}: XLSX={xlsx_val} vs CSV={csv_val}")
            except:
                # String comparison
                xlsx_str = str(xlsx_val).strip()
                csv_str = str(csv_val).strip()
                if xlsx_str != csv_str and xlsx_str.replace('\n', ' ') != csv_str:
                    pass  # Header diffs are ok
        elif xlsx_val is not None and csv_val in (None, ''):
            mismatches.append(f"Row {row_idx}, Col {col_idx}: XLSX={xlsx_val} but CSV=empty")
        elif xlsx_val is None and csv_val not in (None, ''):
            mismatches.append(f"Row {row_idx}, Col {col_idx}: XLSX=empty but CSV={csv_val}")

print(f"\nXLSX vs CSV NUMERIC MISMATCHES: {len(mismatches)}")
for m in mismatches[:30]:
    print(f"  {m}")
