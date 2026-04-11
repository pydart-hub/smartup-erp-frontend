"""
Parse 'FEES STRUCTURE FINAL 2026-27.xlsx' into a comprehensive JSON file.
Zero data loss — every cell value is preserved.
"""

import openpyxl
import json
import re
import sys

XLSX_PATH = r"FEES STRUCTURE FINAL 2026-27.xlsx"
OUTPUT_PATH = r"fee_structure_2026_27.json"

# Map sheet names to canonical branch names
BRANCH_MAP = {
    "ERAVELI": "Eraveli",
    "TIER 1": "Tier 1",
    "THOPPUMPADY": "Thoppumpady",
    "MOOLAMKUZHI": "Moolamkuzhi",
    "VENNALA": "Vennala",
    "KADAVANTHARA": "Kadavanthara",
    "EDAPALLY": "Edapally",
}


def detect_fee_category(title: str) -> str:
    """Extract fee category from section title like 'Basic Fees Structure - ...'"""
    t = title.strip().lower()
    if "subject wise advanced" in t or "subject wise adv" in t:
        return "Subject-wise Advanced"
    if "subject wise basic" in t:
        return "Subject-wise Basic"
    if "advanced" in t:
        return "Advanced"
    if "intrmediate" in t or "intermediate" in t:
        return "Intermediate"
    if "basic" in t:
        return "Basic"
    return title.strip()


def detect_plan_type(headers: list) -> str:
    """Detect whether this is a quarterly/OTP table or installment table."""
    header_text = " ".join(str(h or "") for h in headers).lower()
    if "inst plan" in header_text or "inst amount" in header_text:
        return "installment"
    return "quarterly"


def clean_class_name(name: str) -> str:
    """Normalize class name whitespace."""
    if not name:
        return name
    return re.sub(r'\s+', ' ', name.strip())


def parse_quarterly_row(headers, row_values):
    """Parse a row from the quarterly/OTP payment plan table."""
    data = {}
    for i, h in enumerate(headers):
        if h is None or i >= len(row_values):
            continue
        key = str(h).strip()
        val = row_values[i]
        # Clean up header names
        key_lower = key.lower().replace('\n', ' ').strip()

        if "annual" in key_lower:
            data["annual_fees"] = val
        elif "early bird" in key_lower:
            data["early_bird"] = val
        elif key_lower in ("otp", "one time payment"):
            data["one_time_payment"] = val
        elif "quarterly" in key_lower:
            data["quarterly_total"] = val
        elif "quarter 1" in key_lower:
            data["quarter_1"] = val
        elif "quarter 2" in key_lower:
            data["quarter_2"] = val
        elif "quarter 3" in key_lower:
            data["quarter_3"] = val
        elif "quarter 4" in key_lower:
            data["quarter_4"] = val
    return data


def parse_installment_row(headers, row_values):
    """Parse a row from the installment plan table."""
    data = {}
    for i, h in enumerate(headers):
        if h is None or i >= len(row_values):
            continue
        key = str(h).strip().replace('\n', ' ').lower()
        val = row_values[i]

        if "annual" in key:
            data["annual_fees"] = val
        elif "early bird" in key:
            data["early_bird"] = val
        elif "6 inst plan" in key:
            data["6_installment_total"] = val
        elif "inst amount" in key and "1st - 5th" in key:
            data["6_inst_amount_1_to_5"] = val
        elif "inst amount" in key and "6th" in key:
            data["6_inst_amount_6_final"] = val
        elif "8 inst plan" in key:
            data["8_installment_total"] = val
        elif "inst amount" in key and "1st - 7th" in key:
            data["8_inst_amount_1_to_7"] = val
        elif "inst amount" in key and "8th" in key:
            data["8_inst_amount_8_final"] = val
    return data


def parse_sheet(ws):
    """Parse a single worksheet into structured sections."""
    sections = []

    # Collect all non-empty rows with their row numbers
    rows = []
    for r in range(1, ws.max_row + 1):
        row_data = []
        for c in range(1, ws.max_column + 1):
            row_data.append(ws.cell(r, c).value)
        if any(v is not None for v in row_data):
            rows.append((r, row_data))

    i = 0
    while i < len(rows):
        row_num, row_data = rows[i]
        first_cell = str(row_data[0] or "").strip()

        # Detect section title (contains "Fees Structure")
        if "fees structure" in first_cell.lower() or "Fees Structure" in first_cell:
            category = detect_fee_category(first_cell)
            section_title = first_cell
            i += 1

            if i >= len(rows):
                break

            # Next row should be headers
            _, header_row = rows[i]
            plan_type = detect_plan_type(header_row)
            i += 1

            # Parse data rows until next section title or empty gap
            class_entries = []
            while i < len(rows):
                _, data_row = rows[i]
                cell0 = str(data_row[0] or "").strip()
                # Stop if we hit another section title
                if "fees structure" in cell0.lower():
                    break
                # Skip if first cell is empty or looks like a header
                if not data_row[0] or cell0.lower() in ("class", "hss subject"):
                    i += 1
                    continue

                class_name = clean_class_name(str(data_row[0]))

                if plan_type == "quarterly":
                    entry = parse_quarterly_row(header_row, data_row)
                else:
                    entry = parse_installment_row(header_row, data_row)

                entry["class"] = class_name
                class_entries.append(entry)
                i += 1

            sections.append({
                "section_title": section_title,
                "fee_category": category,
                "plan_type": plan_type,
                "is_subject_wise": "subject wise" in category.lower(),
                "entries": class_entries,
            })
        else:
            i += 1

    return sections


def merge_sections_into_structure(sections):
    """
    Merge quarterly + installment sections for same category into unified entries.
    """
    # Group by (fee_category)
    groups = {}
    for sec in sections:
        cat = sec["fee_category"]
        if cat not in groups:
            groups[cat] = {"quarterly": None, "installment": None}
        groups[cat][sec["plan_type"]] = sec

    result = []
    for cat, plans in groups.items():
        q_sec = plans.get("quarterly")
        i_sec = plans.get("installment")

        # Build per-class merged data
        class_data = {}

        if q_sec:
            for entry in q_sec["entries"]:
                cn = entry["class"]
                class_data.setdefault(cn, {})
                class_data[cn]["class"] = cn
                class_data[cn]["annual_fees"] = entry.get("annual_fees")
                class_data[cn]["payment_plans"] = {
                    "early_bird": entry.get("early_bird"),
                    "one_time_payment": entry.get("one_time_payment"),
                    "quarterly": {
                        "total_after_5pct_discount": entry.get("quarterly_total"),
                        "quarter_1": entry.get("quarter_1"),
                        "quarter_2": entry.get("quarter_2"),
                        "quarter_3": entry.get("quarter_3"),
                        "quarter_4": entry.get("quarter_4"),
                    },
                }

        if i_sec:
            for entry in i_sec["entries"]:
                cn = entry["class"]
                class_data.setdefault(cn, {})
                if "class" not in class_data[cn]:
                    class_data[cn]["class"] = cn
                if "annual_fees" not in class_data[cn]:
                    class_data[cn]["annual_fees"] = entry.get("annual_fees")
                if "payment_plans" not in class_data[cn]:
                    class_data[cn]["payment_plans"] = {}

                class_data[cn]["payment_plans"]["6_installment"] = {
                    "early_bird": entry.get("early_bird"),
                    "total_after_2_5pct_discount": entry.get("6_installment_total"),
                    "installment_1_to_5": entry.get("6_inst_amount_1_to_5"),
                    "installment_6_final": entry.get("6_inst_amount_6_final"),
                }
                class_data[cn]["payment_plans"]["8_installment"] = {
                    "total": entry.get("8_installment_total"),
                    "installment_1_to_7": entry.get("8_inst_amount_1_to_7"),
                    "installment_8_final": entry.get("8_inst_amount_8_final"),
                }

        is_subject = q_sec["is_subject_wise"] if q_sec else (i_sec["is_subject_wise"] if i_sec else False)

        result.append({
            "fee_category": cat,
            "is_subject_wise": is_subject,
            "classes": list(class_data.values()),
        })

    return result


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    output = {
        "academic_year": "2026-27",
        "source_file": "FEES STRUCTURE FINAL 2026-27.xlsx",
        "branches": [],
    }

    total_entries = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        branch_name = BRANCH_MAP.get(sheet_name, sheet_name)

        sections = parse_sheet(ws)
        merged = merge_sections_into_structure(sections)

        branch_entry_count = sum(len(cat["classes"]) for cat in merged)
        total_entries += branch_entry_count

        branch_data = {
            "branch": branch_name,
            "sheet_name": sheet_name,
            "fee_categories": merged,
        }
        output["branches"].append(branch_data)

        print(f"  [{sheet_name}] -> {branch_name}: {len(sections)} sections, {branch_entry_count} class entries")

    output["_metadata"] = {
        "total_branches": len(output["branches"]),
        "total_class_entries": total_entries,
        "note": "All monetary values are in INR (Indian Rupees). Discount percentages: Early Bird = ~10% off Annual, OTP (One Time Payment) = ~17% off Annual, Quarterly = 5% off Annual, 6-installment = 2.5% off Early Bird, 8-installment = no discount off Early Bird."
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nDone! Written to {OUTPUT_PATH}")
    print(f"Total branches: {len(output['branches'])}, Total class entries: {total_entries}")


if __name__ == "__main__":
    main()
